import openpyxl

from openpyxl.cell import Cell
from openpyxl.worksheet.worksheet import Worksheet


class ProcessXlsx:
    """
    用 openpyxl 处理 xlsx 格式 excel
    """

    def __init__(self, filename=None):
        self.__filename = filename
        self.__column_dict = None  # 列的映射字典表, 可以自定义, 格式参考下方的 setter 方法

    @property
    def column_dict(self):
        return self.__column_dict

    @column_dict.setter
    def column_dict(self, _column_dict):
        """
        _column_dict = {
            column_name1: 'A'
            column_name2: 'B'
            ...
        }
        """
        self.__column_dict = _column_dict

    def work_book(self) -> openpyxl.workbook.workbook.Workbook:
        return openpyxl.load_workbook(self.__filename)

    def work_sheets_by_number(self, sheet_number=0) -> openpyxl.worksheet.worksheet.Worksheet:
        return self.work_book()[self.work_book().sheetnames[sheet_number]]

    def work_sheets_by_name(self, sheet_name="Sheet1") -> openpyxl.worksheet.worksheet.Worksheet:
        return self.work_book()[sheet_name]

    @staticmethod
    def max_row(work_sheet: Worksheet) -> int:
        return work_sheet.max_row

    def rows(self, row_number=1, sheet_number=0) -> list:
        """
        获取该行的所有值
        """
        _values = []
        ws = self.work_sheets_by_number(sheet_number)
        for cell in ws[row_number]:
            cell: Cell
            _values.append(cell.value)
        return _values

    @staticmethod
    def max_column(work_sheet: Worksheet) -> int:
        return work_sheet.max_column

    def columns(self, column_name, _column_dict, sheet_number=0) -> list:
        """
        获取该列的所有值
        """
        _values = []
        if _column_dict is None:
            ws = self.work_sheets_by_number(sheet_number)
            for value in ws[column_name]:
                _values.append(value)

        ws = self.work_sheets_by_number(sheet_number)
        for value in ws[self.column_dict.get(column_name)]:
            _values.append(value)
        return _values
